#!/usr/bin/env python3
"""
Build a static HTML preview: PlantUML (SVG) + StatusCode markdown + Excel Sheet1 table.

Usage (from repo root or any cwd):
  pip install -r docs/observable/kv-client-excel/preview/requirements-preview.txt
  python3 docs/observable/kv-client-excel/preview/build_preview.py

Requires Java + plantuml.jar (set PLANTUML_JAR or uses /tmp/plantuml.jar if present).
Open preview/dist/index.html in a browser (file:// is OK for local preview).
"""

from __future__ import annotations

import html
import os
import re
import shutil
import subprocess
import sys
import urllib.request
from pathlib import Path

PREVIEW_DIR = Path(__file__).resolve().parent
KV_EXCEL_DIR = PREVIEW_DIR.parent
DIST_DIR = PREVIEW_DIR / "dist"
DIAGRAM_DIR = DIST_DIR / "diagrams"
VIBE_ROOT = KV_EXCEL_DIR.parent.parent.parent

XLSX_PATH = KV_EXCEL_DIR / "kv-client-观测-调用链与URMA-TCP.xlsx"
STATUS_CODES_MD = VIBE_ROOT / "docs/reliability/00-kv-client-visible-status-codes.md"
PUML_DIR = KV_EXCEL_DIR / "puml"
PLANTUML_JAR_CANDIDATES = [
    Path(os.environ.get("PLANTUML_JAR", "")) if os.environ.get("PLANTUML_JAR") else None,
    Path("/tmp/plantuml.jar"),
    PREVIEW_DIR / ".cache" / "plantuml.jar",
]
PLANTUML_DOWNLOAD = "https://github.com/plantuml/plantuml/releases/download/v1.2024.7/plantuml-1.2024.7.jar"

SHEET1_TITLE_CANDIDATES = (
    "Sheet1_调用链路分析",
    "Sheet1_调用链路",
    "Sheet1",
)


def ensure_plantuml_jar() -> Path | None:
    for p in PLANTUML_JAR_CANDIDATES:
        if p and p.is_file():
            return p
    cache = PREVIEW_DIR / ".cache"
    cache.mkdir(parents=True, exist_ok=True)
    dest = cache / "plantuml.jar"
    if not dest.is_file():
        print("Downloading plantuml.jar …", file=sys.stderr)
        urllib.request.urlretrieve(PLANTUML_DOWNLOAD, dest)
    return dest if dest.is_file() else None


def _puml_diagram_id(puml_path: Path) -> str:
    """PlantUML names output `<id>.svg` from first line: @startuml <id>."""
    first = puml_path.read_text(encoding="utf-8").splitlines()[:3]
    for line in first:
        line = line.strip()
        if line.startswith("@startuml"):
            rest = line[len("@startuml") :].strip()
            return rest.split()[0] if rest else puml_path.stem
    return puml_path.stem


def render_plantuml(jar: Path, puml_files: list[Path]) -> list[tuple[str, str]]:
    """Returns list of (display_title, relative_svg_path)."""
    DIAGRAM_DIR.mkdir(parents=True, exist_ok=True)
    out: list[tuple[str, str]] = []
    for puml in sorted(puml_files):
        if puml.name.startswith("."):
            continue
        subprocess.run(
            [
                "java",
                "-jar",
                str(jar),
                "-tsvg",
                "-o",
                str(DIAGRAM_DIR),
                str(puml),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        did = _puml_diagram_id(puml)
        svg_name = f"{did}.svg"
        svg_path = DIAGRAM_DIR / svg_name
        if svg_path.is_file():
            # 人读标题：保留原 puml 文件名（无扩展名）为主，括号内为 diagram id
            out.append((puml.stem, f"diagrams/{svg_name}"))
        else:
            print(f"warn: missing SVG for {puml.name} (expected {svg_name})", file=sys.stderr)
    return out


def sheet1_to_html_rows() -> tuple[str, str]:
    """Returns (title, html_table_or_message)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return (
            "Excel Sheet1",
            "<p class=\"warn\">未安装 openpyxl。请执行: <code>pip install -r preview/requirements-preview.txt</code></p>",
        )

    if not XLSX_PATH.is_file():
        return (
            "Excel Sheet1",
            f"<p class=\"warn\">未找到工作簿: <code>{html.escape(str(XLSX_PATH))}</code>。<br>"
            "请先在 <code>kv-client-excel/scripts/</code> 下运行生成脚本产出 xlsx，或从交付包拷贝该文件到同目录。</p>",
        )

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = None
    title = ""
    for name in SHEET1_TITLE_CANDIDATES:
        if name in wb.sheetnames:
            ws = wb[name]
            title = name
            break
    if ws is None:
        ws = wb.worksheets[0]
        title = ws.title
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        wb.close()
        return title, "<p>（空表）</p>"
    ths = "".join(f"<th>{html.escape(str(h) if h is not None else '')}</th>" for h in headers)
    body_rows = []
    for row in rows_iter:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = []
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else None
            cells.append(f"<td>{html.escape(str(v) if v is not None else '')}</td>")
        body_rows.append("<tr>" + "".join(cells) + "</tr>")
    wb.close()
    table = (
        f'<div class="table-wrap"><table class="sheet"><thead><tr>{ths}</tr></thead>'
        f"<tbody>{''.join(body_rows)}</tbody></table></div>"
    )
    return title, table


def markdown_to_html(md_path: Path) -> str:
    try:
        import markdown as md_lib
    except ImportError:
        raw = md_path.read_text(encoding="utf-8")
        return f"<pre class=\"fallback-md\">{html.escape(raw)}</pre>"

    text = md_path.read_text(encoding="utf-8")
    # Strip repo-specific top links noise slightly for preview
    return md_lib.markdown(
        text,
        extensions=["tables", "fenced_code", "nl2br", "sane_lists"],
        extension_configs={"fenced_code": {"lang_prefix": "language-"}},
    )


def build_index(diagrams: list[tuple[str, str]], sheet_title: str, sheet_html: str, body_html: str) -> str:
    nav_items = [
        ("#diagrams", "PlantUML 图"),
        ("#status-codes", "StatusCode 速查"),
        ("#sheet1", f"调用链表（{html.escape(sheet_title)}）"),
    ]
    nav = "<ul>" + "".join(f'<li><a href="{href}">{label}</a></li>' for href, label in nav_items) + "</ul>"
    figures = []
    for stem, rel in diagrams:
        cap = stem.replace("_", " ")
        figures.append(
            f'<figure id="fig-{html.escape(stem, quote=True)}">'
            f'<h4>{html.escape(cap)}</h4>'
            f'<object type="image/svg+xml" data="{html.escape(rel)}" class="diagram">'
            f'<img src="{html.escape(rel)}" alt="{html.escape(cap)}"/></object></figure>'
        )
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>KV Client 定位定界 — 预览</title>
  <style>
    :root {{
      --bg: #f6f7f9;
      --card: #fff;
      --text: #1a1a1a;
      --muted: #5c5c5c;
      --border: #d8dce3;
      --accent: #0b57d0;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      line-height: 1.55;
      color: var(--text);
      background: var(--bg);
      margin: 0;
      padding: 1rem 1.25rem 3rem;
      max-width: 1200px;
      margin-left: auto;
      margin-right: auto;
    }}
    header {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 1rem 1.25rem;
      margin-bottom: 1.25rem;
    }}
    header h1 {{ margin: 0 0 0.35rem 0; font-size: 1.35rem; }}
    header p {{ margin: 0; color: var(--muted); font-size: 0.9rem; }}
    nav ul {{ margin: 0.5rem 0 0 0; padding-left: 1.2rem; }}
    section {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 1rem 1.25rem;
      margin-bottom: 1.25rem;
    }}
    section h2 {{
      margin-top: 0;
      font-size: 1.15rem;
      border-bottom: 1px solid var(--border);
      padding-bottom: 0.35rem;
    }}
    .diagram {{ width: 100%; max-height: 85vh; }}
    figure {{ margin: 1rem 0; }}
    figure h4 {{ margin: 0 0 0.5rem 0; font-size: 0.95rem; color: var(--muted); }}
    .table-wrap {{ overflow-x: auto; max-height: 70vh; overflow-y: auto; }}
    table.sheet {{
      border-collapse: collapse;
      font-size: 0.78rem;
      width: max-content;
      min-width: 100%;
    }}
    table.sheet th, table.sheet td {{
      border: 1px solid var(--border);
      padding: 0.35rem 0.5rem;
      vertical-align: top;
      max-width: 28rem;
      white-space: normal;
      word-break: break-word;
    }}
    table.sheet th {{
      background: #eef1f6;
      position: sticky;
      top: 0;
      z-index: 1;
    }}
    .warn {{ color: #8a4b00; background: #fff8e6; padding: 0.75rem; border-radius: 8px; }}
    #status-codes .markdown-body table {{ border-collapse: collapse; width: 100%; font-size: 0.85rem; }}
    #status-codes .markdown-body th, #status-codes .markdown-body td {{
      border: 1px solid var(--border); padding: 0.35rem 0.45rem;
    }}
    #status-codes .markdown-body pre {{
      overflow-x: auto;
      background: #f0f2f5;
      padding: 0.75rem;
      border-radius: 8px;
      font-size: 0.82rem;
    }}
    #status-codes .markdown-body code {{
      background: #f0f2f5;
      padding: 0.1rem 0.3rem;
      border-radius: 4px;
      font-size: 0.88em;
    }}
    .fallback-md {{ white-space: pre-wrap; font-size: 0.85rem; }}
    a {{ color: var(--accent); }}
  </style>
</head>
<body>
  <header>
    <h1>KV Client 定位定界 — 本地预览</h1>
    <p>由 <code>preview/build_preview.py</code> 生成；含 PlantUML SVG、可见 StatusCode 文档、Excel 第一工作表。材料路径以本仓库为准。</p>
    <nav>{nav}</nav>
  </header>

  <section id="diagrams">
    <h2>PlantUML 图（puml/）</h2>
    {chr(10).join(figures) if figures else '<p class="warn">未生成任何图（检查 Java / plantuml.jar）。</p>'}
  </section>

  <section id="status-codes">
    <h2>关键错误码文档</h2>
    <p class="muted" style="color:var(--muted);font-size:0.9rem;">来源: <code>docs/reliability/00-kv-client-visible-status-codes.md</code></p>
    <div class="markdown-body">{body_html}</div>
  </section>

  <section id="sheet1">
    <h2>观测 Excel — 调用链（第一工作表）</h2>
    <p class="muted" style="color:var(--muted);font-size:0.9rem;">工作表: <strong>{html.escape(sheet_title)}</strong> · 文件: <code>kv-client-观测-调用链与URMA-TCP.xlsx</code></p>
    {sheet_html}
  </section>
</body>
</html>
"""


def main() -> int:
    if not STATUS_CODES_MD.is_file():
        print(f"error: missing {STATUS_CODES_MD}", file=sys.stderr)
        return 1

    DIST_DIR.mkdir(parents=True, exist_ok=True)
    if DIAGRAM_DIR.exists():
        shutil.rmtree(DIAGRAM_DIR)
    DIAGRAM_DIR.mkdir(parents=True, exist_ok=True)

    puml_files = list(PUML_DIR.glob("*.puml"))
    diagrams: list[tuple[str, str]] = []
    jar = ensure_plantuml_jar()
    if jar and puml_files:
        try:
            diagrams = render_plantuml(jar, puml_files)
        except (subprocess.CalledProcessError, FileNotFoundError) as e:
            print(f"warn: PlantUML failed ({e}); diagrams section will be empty.", file=sys.stderr)
    elif not jar:
        print("warn: no plantuml.jar; skip diagrams.", file=sys.stderr)

    sheet_title, sheet_html = sheet1_to_html_rows()
    body_html = markdown_to_html(STATUS_CODES_MD)
    index_html = build_index(diagrams, sheet_title, sheet_html, body_html)
    (DIST_DIR / "index.html").write_text(index_html, encoding="utf-8")
    print(f"Wrote {DIST_DIR / 'index.html'}")
    print(f"Open in browser: file://{(DIST_DIR / 'index.html').resolve()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
