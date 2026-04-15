#!/usr/bin/env python3
"""Generate KV Client observability workbook for fault triage."""

import re
from pathlib import Path

from openpyxl import Workbook

from sheet1_system_presets import SHEET1_URMA_OS, merge_exclusive
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[3]
OUT = REPO_ROOT / "docs" / "observable" / "workbook" / "kv-client" / "kv-client-观测-调用链与URMA-TCP.xlsx"
SOURCE_ROOT = Path("/home/t14s/workspace/git-repos/yuanrong-datasystem")

# 责任归属取值约定：用户参数 | OS | URMA | 数据系统逻辑 | RPC框架
CHAIN_HEADERS = [
    "接口",
    "调用链树（正向·根→叶）\n+ 调用逻辑 + 故障预期 + case/位置",
    "责任归属",
    "错误返回值与描述",
    "URMA接口调用\n(ds_urma_* / UMDK)",
    "URMA错误码与完整日志线索\n（与右侧 OS 列互斥）",
    "OS接口调用\n(syscall / libzmq)",
    "OS错误码与完整日志线索\n（与左侧 URMA 列互斥）",
    "开发定位步骤\n(grep / 日志 / 下一步)",
    "备注(返回/重试/代码锚点)",
]

# 每行 11 元组为旧格式（默认可失败）；12 元组最后一项为【故障预期】：无错误 | 可失败（…）| 可失败-降级 等
CHAIN_ROWS = [
    (
        "Init",
        "KVClient::Init|_ObjectClientImpl::Init|_InitClientWorkerConnect|_ClientWorkerRemoteApi::Init|_Connect: GetSocketPath+RegisterClient|_PostRegisterClient|_FastTransportHandshake(可选UB)|_MmapManager+ListenWorker+线程收尾",
        "主路径：无参数/网络/资源异常时（与 §1.1 对齐）",
        "client1（Init 主路径·含与入口 Worker Register）",
        "数据系统+RPC",
        "ZMQ 控制面 + 可选 ds_urma_*",
        "OK",
        "正常返回",
        "无",
        "与步骤1 Init puml、Markdown §1.1/§2.1 对照；本行不描述具体故障点",
        "kv_client.cpp; object_client_impl.cpp; client_worker_common_api.cpp",
        "无错误",
    ),
    ("Init", "KVClient::Init|_ObjectClientImpl::Init", "HostPort/凭证非法", "client1", "用户参数", "-", "K_INVALID", "直接RETURN", "无", "核对 IP/端口/AKSK 与部署一致", "object_client_impl.cpp Init"),
    ("Init", "InitClientWorkerConnect|_Connect|_RegisterClient", "Register 连不上/超时", "client1->worker1（SDK到入口Worker控制面RPC）", "RPC框架", "socket/connect + ZMQ 请求应答", "1002/1001; Register client failed", "Status返回", "一般无", "查 worker 监听、防火墙、ZMQ 会话是否复用异常", "client_worker_common_api.cpp Connect"),
    ("Init", "InitClientWorkerConnect|_Connect|_RegisterClient", "Register 后 socket 读写失败/超时", "client1->worker1（SDK到入口Worker控制面RPC）", "OS+RPC框架", "send/recv + zmq_send/zmq_recv + zmq_poll", "1002/1001/19; timeout/unavailable/try_again", "Status返回", "视阶段重试", "查 ZMQ HWM/超时参数、网络抖动、对端是否卡死", "client_worker_common_api.cpp; client_worker_remote_api.cpp"),
    ("Init", "PostRegisterClient|_FastTransportHandshake|_UrmaInit", "UB 握手 urma_init 等失败", "client1", "URMA", "urma_init / urma_create_* / urma_query_device", "1004; Failed to urma init", "LOG_IF_ERROR不阻断Init", "无", "查 UMDK、设备名、bonding、驱动", "urma_manager.cpp UrmaInit"),
    ("Init", "PostRegisterClient|_InitMemoryBufferPool", "UB 客户端池 mmap 失败", "client1", "OS", "mmap(MAP_ANON,-1)", "K_OUT_OF_MEMORY(6); Failed to allocate memory buffer pool", "链上RETURN", "无", "查进程内存限制、vm.max_map_count、ulimit", "urma_manager.cpp InitMemoryBufferPool"),
    ("Init", "PostRegisterClient|_ExchangeJfr|_ImportRemoteJfr", "ExchangeJfr/import 失败", "client1->worker1（SDK与入口Worker UB握手）", "URMA", "urma_import_jfr / urma_advise_jfr", "Fast transport handshake failed…fall back", "仅ERROR日志", "无", "查 client-worker UB 握手、对端 jfr 可达性", "urma_manager.cpp ExchangeJfr"),
    ("Init", "PostRegisterClient|_RecvPageFd|_SockRecvFd", "UDS 收 fd 失败", "client1", "OS", "recvmsg(SCM_RIGHTS) / sendmsg", "K_UNKNOWN_ERROR; Pass fd…/invalid fd", "线程内失败", "无", "查 UDS 路径、fd 上限、对端是否先发 fd", "fd_pass.cpp SockRecvFd"),
    (
        "MGet/Get",
        "ObjectClientImpl::Get|_GetBuffersFromWorker|_ClientWorkerRemoteApi::Get|_PreGet|_PrepareUrmaBuffer(UB可选)|_stub_->Get→W1|_ProcessGetObjectRequest|_TryGetObjectFromLocal|_[未全命中→]_TryGetObjectFromRemote|_ProcessObjectsNotExistInLocal|_QueryMetadataFromMaster→W2|_GetObjectsFromAnywhere→W3@queryMeta.address|_ReturnToClient(last_rc)|_Client ProcessGetResponse|_FillUrmaBuffer(UB可选)",
        "主路径：无本表下列各「case」行所述异常时（与 §1.5、步骤2 puml 对齐）",
        "client→W1→(W2)→(W3)→client",
        "数据系统",
        "ZMQ + Directory(QueryMeta) + Worker↔Worker + 可选URMA",
        "OK / last_rc OK",
        "正常返回",
        "无",
        "本地命中时 W2/W3 可不发生；跨节点读时顺序为 问对象目录(Directory)→按地址拉副本",
        "object_client_impl.cpp; client_worker_remote_api.cpp; worker_oc_service_get_impl.cpp",
        "无错误",
    ),
    ("MGet/Get", "ObjectClientImpl::Get|_ClientWorkerRemoteApi::Get|_PreGet", "subTimeout/batch/offset 非法", "client1", "用户参数", "-", "K_INVALID; subTimeoutMs out of range", "直接RETURN", "无", "核对 SDK 入参与 OBJECT_KEYS_MAX_LIMIT", "client_worker_base_api.cpp PreGet"),
    ("MGet/Get", "ClientWorkerRemoteApi::Get|_PrepareUrmaBuffer", "UB 缓冲准备失败降级", "client1", "URMA(可降级)", "GetMemoryBufferHandle/GetMemoryBufferInfo(池)", "WARNING; fallback to TCP/IP payload", "不返回URMA码继续RPC", "无", "功能可成功；性能差则查 UB 池初始化与并发", "client_worker_base_api.cpp PrepareUrmaBuffer"),
    ("MGet/Get", "ClientWorkerRemoteApi::Get|_RetryOnError|_stub_->Get", "控制面 RPC 失败/重试耗尽", "client1->worker1（SDK到入口Worker控制面RPC）", "RPC框架", "stub_->Get(ZMQ) + zmq_poll超时", "1001/1002/19; RPC Retry detail", "Status返回", "RetryOnError", "无 worker 入口日志→网络；有入口→worker 慢或下游", "client_worker_remote_api.cpp Get"),
    ("MGet/Get", "ClientWorkerRemoteApi::Get|_RetryOnError|_stub_->Get", "RPC socket 读写失败/超时", "client1->worker1（SDK到入口Worker控制面RPC）", "OS+RPC框架", "send/recv + zmq_send/zmq_recv + zmq_poll", "1001/1002/19; timeout/unavailable/try_again", "重试后返回", "RetryOnError", "抓包+对齐 SDK/worker 同 key 时间线", "client_worker_remote_api.cpp"),
    ("MGet/Get", "Get lambda|_last_rc 评估", "业务 last_rc 超时/OOM 触发重试", "client1->worker1（SDK到入口Worker控制面RPC）", "数据系统逻辑", "-", "last_rc 1001/19/6+全失败", "重试", "见RPC", "查 worker 负载、远端拉取、内存", "client_worker_remote_api.cpp Get lambda"),
    ("MGet/Get", "ProcessGetObjectRequest|_QueryMeta", "etcd 续约不可用", "worker1（入口Worker本地处理）", "数据系统逻辑", "etcd RawGet(经封装)", "1002; etcd is unavailable", "last_rc", "worker内策略", "查 etcd 集群与 worker 续约", "worker_oc_service_get_impl.cpp"),
    ("MGet/Get", "ProcessGetObjectRequest|_QueryMeta", "QueryMeta/Directory 失败", "worker1->worker2（对象目录分片）", "数据系统逻辑+RPC", "gRPC/RPC 到 Directory Worker（hash ring）", "K_RUNTIME_ERROR; Query from master failed", "last_rc", "视实现", "查 Directory 分片可用性与目录一致性（日志原文可能仍为 Query from master failed）", "worker_oc_service_get_impl.cpp"),
    ("MGet/Get", "ProcessGetObjectRequest|_GetObjectFromRemote", "远端 Worker 拉取失败", "worker1->worker3（跨节点数据拉取请求）", "数据系统逻辑+RPC", "Worker↔Worker stub", "Get from remote failed; 1002", "last_rc/RPC", "TryReconnect…", "对齐 worker3 日志与同 key", "worker_oc_service_get_impl.cpp"),
    ("MGet/Get", "UrmaWritePayload|_urma_write/read", "URMA write/read 失败", "worker3（数据副本Worker本地数据面）", "URMA", "urma_write / urma_read", "K_RUNTIME_ERROR; Failed to urma write/read", "汇总回 client", "-", "查对端可达、完成错误、UMDK 日志", "urma_manager.cpp"),
    ("MGet/Get", "PollJfcWait|_urma_wait/poll/rearm_jfc", "URMA poll/wait jfc 失败", "worker3（数据副本Worker本地数据面）", "URMA", "urma_poll_jfc / urma_wait_jfc / urma_rearm_jfc", "1004; Failed to poll/wait jfc", "汇总回 client", "-", "查 CQ、事件模式、设备状态", "urma_manager.cpp PollJfcWait"),
    ("MGet/Get", "CheckUrmaConnectionStable", "URMA 需重连", "worker1/worker3（入口或数据面连通性检测）", "URMA", "CheckUrmaConnectionStable(逻辑)", "1006; need to reconnect", "Status", "-", "重建 UB 会话后再试", "urma_manager.cpp"),
    ("MGet/Get", "FillUrmaBuffer|_GetBuffersFromWorker", "FillUrmaBuffer 越界/组装失败", "client1", "数据系统逻辑", "-", "K_RUNTIME_ERROR(5); UB payload overflow", "直接RETURN", "无", "查 worker 返回的 payload 与 UB 尺寸一致性", "client_worker_base_api.cpp FillUrmaBuffer"),
    ("MGet/Get", "MmapShmUnit|_LookupUnitsAndMmapFd", "SHM mmap 失败", "client1", "OS+数据系统", "mmap(经 LookupUnitsAndMmapFd)", "Get mmap entry failed", "单key失败聚合", "无", "查 fd 是否有效、是否同机、表项是否注册", "object_client_impl.cpp MmapShmUnit"),
    ("MGet/Get", "ProcessGetObjectRequest", "K_NOT_FOUND 对象不存在", "worker1/worker3（入口到数据副本）", "数据系统逻辑", "-", "K_NOT_FOUND", "last_rc", "-", "数据/副本/过期，非 URMA/OS 根因", "worker_oc_service_get_impl.cpp"),
    (
        "MSet/Put",
        "ObjectClientImpl::MSet|_ClientWorkerRemoteApi::MultiPublish|_stub_->MultiPublish→W1|_W1 写处理(目录+副本; Directory 调用多在 W1 内)|_W3 数据面落盘或 URMA(与步骤3 puml 一致)|_响应→Client|_HandleShmRefCountAfterMultiPublish",
        "主路径：无 RPC/scaling/OOM/UB 等异常时",
        "client→W1(内部可达 W3)",
        "数据系统",
        "ZMQ MultiPublish + 内部副本链",
        "OK",
        "正常返回",
        "无",
        "写路径 Excel/步骤3 未单画 W2；目录写入在 W1 逻辑内完成",
        "object_client_impl.cpp; client_worker_remote_api.cpp",
        "无错误",
    ),
    ("MSet/Put", "SendBufferViaUb|_UrmaWritePayload", "UB 单对象发送失败", "client1", "URMA", "UrmaWritePayload→urma_write 链", "K_INVALID; Failed to send buffer via UB", "可改走非UB", "无", "查 UB 池大小与对象总大小", "client_worker_base_api.cpp SendBufferViaUb"),
    ("MSet/Put", "Publish|_RetryOnError|_stub_->Publish", "Publish RPC 失败", "client1->worker1（SDK到入口Worker控制面RPC）", "RPC框架", "stub_->Publish + zmq_poll超时", "1001/1002/19", "Status", "RetryOnError", "同 Get 控制面", "client_worker_remote_api.cpp Publish"),
    ("MSet/Put", "Publish/MultiPublish RPC", "Publish/MultiPublish socket 读写失败/超时", "client1->worker1（SDK到入口Worker控制面RPC）", "OS+RPC框架", "send/recv + zmq_send/zmq_recv + zmq_poll", "1001/1002/19", "重试后返回", "RetryOnError", "查 ZMQ 超时与水位，关注队列堆积", "client_worker_remote_api.cpp"),
    ("MSet/Put", "MultiPublish|_RetryOnError", "MultiPublish 扩缩容/内存", "client1->worker1（SDK到入口Worker控制面RPC）", "数据系统逻辑", "-", "32 K_SCALING; 6 OOM", "Status", "RetryOnError含扩展码", "查集群 scaling 状态与批次大小", "client_worker_remote_api.cpp MultiPublish"),
]

assert len(SHEET1_URMA_OS) == len(CHAIN_ROWS), "sheet1_system_presets.SHEET1_URMA_OS 必须与 CHAIN_ROWS 行数一致"

OS_HEADERS = [
    "OS接口/syscall",
    "对应接口(Init/Get/Set)",
    "关键日志(检测用)",
    "映射关系(数据系统错误码)",
    "错误码(分行+含义)",
    "可能原因与应对",
]

OS_ROWS = [
    (
        "socket/connect",
        "Init",
        "Register client failed\nrpc unavailable",
        "ETIMEDOUT/ECONNREFUSED -> K_RPC_DEADLINE_EXCEEDED(1001)/K_RPC_UNAVAILABLE(1002)",
        "ETIMEDOUT(110): 连接超时\nECONNREFUSED(111): 对端拒绝\nENETUNREACH(101): 网络不可达",
        "可能: worker未监听/网络路由异常；应对: 查端口、防火墙、路由、worker存活",
    ),
    (
        "send/recv (含 zmq_send/zmq_recv)",
        "Get/Set",
        "rpc unavailable\ntry again\nsend/recv failed",
        "EPIPE/ECONNRESET/EAGAIN -> K_RPC_UNAVAILABLE(1002)/K_TRY_AGAIN(19)",
        "EPIPE(32): 连接已断开\nECONNRESET(104): 对端重置\nEAGAIN(11): 资源暂不可用",
        "可能: 网络抖动/对端重启/连接复用异常；应对: 对齐client-worker同key时间线并抓包",
    ),
    (
        "zmq_poll",
        "Init/Get/Set",
        "rpc timeout\ntime elapsed",
        "ETIMEDOUT -> K_RPC_DEADLINE_EXCEEDED(1001)",
        "ETIMEDOUT(110): 等待超时",
        "可能: 队列堆积/线程池阻塞/超时配置过小；应对: 查超时参数和worker负载",
    ),
    (
        "sendmsg/recvmsg (UDS传fd)",
        "Init",
        "Pass fd meets unexpected error\nUnexpected EOF read\ninvalid fd",
        "EBADF/EOF -> K_UNKNOWN_ERROR 或 K_RUNTIME_ERROR(5)",
        "EBADF(9): 非法文件描述符\nEINTR(4): 调用被中断\nEOF: 对端提前关闭",
        "可能: fd传递时序异常/对端提前退出；应对: 联查SockSendFd/SockRecvFd配对日志",
    ),
    (
        "mmap (anon/file)",
        "Init/Get",
        "Failed to allocate memory buffer pool\nGet mmap entry failed",
        "ENOMEM/EBADF/EACCES -> K_OUT_OF_MEMORY(6)/K_RUNTIME_ERROR(5)",
        "ENOMEM(12): 内存不足\nEBADF(9): fd无效\nEACCES(13): 权限不足",
        "可能: 内存上限/fd异常/权限问题；应对: 查ulimit、vm.max_map_count、fd来源",
    ),
]

URMA_HEADERS = [
    "URMA接口",
    "对应接口(Init/Get/Set)",
    "关键日志(检测用)",
    "映射关系(数据系统错误码)",
    "错误码(分行+含义)",
    "可能原因与应对",
]

URMA_ROWS = [
    (
        "urma_init / urma_uninit",
        "Init",
        "Failed to urma init, ret = %d\nFailed to urma uninit, ret = %d",
        "ret != URMA_SUCCESS -> K_URMA_ERROR(1004)",
        "URMA_EINVAL(22): 参数非法\nURMA_EEXIST(17): 重复初始化\nURMA_FAIL(0x1000): 通用失败",
        "可能: UMDK/驱动/设备配置异常；应对: 查设备、驱动、UMDK部署与权限",
    ),
    (
        "urma_get_device_by_name / urma_get_eid_list / urma_query_device",
        "Init",
        "Failed to urma get device by name, errno = %d\nFailed to urma get eid list, errno = %d",
        "接口失败 -> K_URMA_ERROR(1004) 或 K_RUNTIME_ERROR(5)",
        "URMA_EINVAL(22): 参数非法\nURMA_ENOMEM(12): 内存不足\nURMA_FAIL(0x1000): 设备查询失败",
        "可能: 设备名不匹配/EID配置异常；应对: 核对 DS_URMA_DEV_NAME、bonding、EID",
    ),
    (
        "urma_create_* / urma_import_jfr / urma_advise_jfr",
        "Init/Get/Set",
        "Failed to urma create ...\nFailed to import target jfr\nFailed to advise jfr",
        "创建/导入失败 -> K_URMA_ERROR(1004) 或 K_RUNTIME_ERROR(5)",
        "URMA_ENOPERM(1): 模式或权限不允许\nURMA_EINPROGRESS(115): 连接处理中\nURMA_FAIL(0x1000): 导入失败",
        "可能: 资源上限/握手时序/对端jfr状态异常；应对: 查连接重建与两端握手日志",
    ),
    (
        "urma_write / urma_read / urma_post_jfs_wr",
        "Get/Set",
        "Failed to urma write object ...\nFailed to urma read object ...",
        "数据面失败 -> K_URMA_ERROR(1004) 或 K_RUNTIME_ERROR(5)",
        "URMA_ETIMEOUT(110): 操作超时\nURMA_EAGAIN(11): 资源暂不可用\nURMA_ENOMEM(12): 资源不足",
        "可能: 链路抖动/远端状态异常/CQ积压；应对: 结合 CR.status 与 jfc 轮询日志",
    ),
    (
        "urma_wait_jfc / urma_poll_jfc / urma_rearm_jfc",
        "Get/Set",
        "Failed to wait jfc, ret = %d\nFailed to poll jfc, ret = %d\nFailed to rearm jfc, status = %d",
        "轮询失败 -> K_URMA_ERROR(1004)",
        "URMA_CR_ACK_TIMEOUT_ERR(9): ACK超时\nURMA_CR_REM_ACCESS_ABORT_ERR(8): 远端访问中止\nURMA_CR_RNR_RETRY_CNT_EXC_ERR(10): RNR重试超限",
        "可能: CQ事件流转异常/对端无buffer/链路不稳；应对: 优先看 CR.status，再看重建JFS是否触发",
    ),
    (
        "CheckUrmaConnectionStable(逻辑)",
        "Get/Set",
        "No existing connection requires creation.\nUrma connect unstable, need to reconnect!",
        "连接不稳定 -> K_URMA_NEED_CONNECT(1006)",
        "K_URMA_NEED_CONNECT(1006): 需要重连（非OS connect错误）",
        "可能: 连接已失效或实例ID不一致；应对: 触发重建UB会话并重试",
    ),
]

PERF_HEADERS = [
    "场景",
    "关键路径阶段",
    "发生位置",
    "热点/等待类型",
    "观测点(代码/指标)",
    "建议采集命令",
    "判定口径",
]

PERF_ROWS = [
    (
        "MGet/Get",
        "client1->worker1 控制面RPC",
        "client1->worker1",
        "RPC排队/网络RTT/重试开销",
        "PerfKey::RPC_CLIENT_GET_OBJECT; RetryOnError次数; status=1001/1002/19",
        "grep 'Start to send rpc to get object' sdk.log",
        "若重试次数高且worker入口日志稀少，优先网络/RPC框架",
    ),
    (
        "MGet/Get",
        "worker1 本地处理",
        "worker1",
        "线程池排队、锁竞争、本地命中比例",
        "PerfKey::WORKER_PROCESS_GET_FROM_LOCAL / _FROM_REMOTE; 'RPC timeout. time elapsed'",
        "grep 'Process Get' worker.log",
        "本地命中低+远端占比高，尾延迟显著上升",
    ),
    (
        "MGet/Get",
        "worker1->worker2 对象目录(Directory)",
        "worker1->worker2",
        "Directory RPC / etcd 租约等待",
        "QueryMeta耗时; etcd unavailable计数",
        "grep -E 'etcd is unavailable|Query from master failed' worker.log",
        "目录链路异常先于数据面修复",
    ),
    (
        "MGet/Get",
        "worker1->worker3 数据拉取",
        "worker1->worker3",
        "跨worker网络+RPC队列",
        "GetObjectFromRemote...耗时; reconnect次数",
        "grep 'Get from remote failed' worker.log",
        "跨机拉取占比高时，先控跨机比例与副本分布",
    ),
    (
        "MGet/Get",
        "worker3 URMA 数据面",
        "worker3",
        "CQ轮询等待、URMA完成错误",
        "urma_wait_jfc / urma_poll_jfc; Failed to poll/wait jfc",
        "grep -E 'poll jfc|wait jfc|urma write|urma read' worker.log",
        "jfc错误增多=URMA链路不稳，可能触发降级",
    ),
    (
        "MGet/Get",
        "client1 UB失败降级TCP",
        "client1",
        "降级导致带宽/CPU拷贝放大",
        "PrepareUrmaBuffer warning次数; payload copy耗时",
        "grep 'fallback to TCP/IP payload' sdk.log",
        "降级率上升且P99升高=优先URMA环境排查",
    ),
    (
        "MSet/Put",
        "Publish/MultiPublish",
        "client1->worker1",
        "重试与批次放大",
        "RetryOnError; K_SCALING/K_OUT_OF_MEMORY",
        "grep 'Send multi publish request error' sdk.log",
        "重试成功但耗时高=容量/扩缩容抖动",
    ),
    (
        "通用",
        "线程切换与调度",
        "worker1/worker3",
        "上下文切换过高",
        "cswch/nvcswch, run queue, futex等待",
        "pidstat -w -p <pid> 1; top -H -p <pid>",
        "自愿/非自愿切换异常升高，优先查锁与线程池配置",
    ),
    (
        "通用",
        "系统调用热点",
        "client1/worker1",
        "recvmsg/sendmsg/mmap/futex阻塞",
        "fd_pass + mmap + RPC框架调用栈",
        "strace -f -tt -T -p <pid> -e trace=network,ipc,memory",
        "syscall时间占比高=优先OS/框架层优化",
    ),
]

# Sheet5: 一行一 case，定界总表（责任归属 + 接口 + 建议）
DEMARCATION_HEADERS = [
    "case编号",
    "责任归属",
    "典型现象(Status或日志片段)",
    "涉及URMA接口",
    "涉及OS接口",
    "优先查模块/位置",
    "定位建议(一句话)",
]

DEMARCATION_ROWS = [
    ("D01", "用户参数", "K_INVALID; key empty / batch超限 / offset越界", "-", "-", "client1 PreGet/CheckValid", "先核对接口文档与入参"),
    ("D02", "RPC框架", "1002/1001/19; Register failed / RPC Retry", "-", "socket/connect(ZMQ)", "client1->worker1 控制面RPC", "无worker入口→网络；有入口→worker或下游慢"),
    ("D02A", "OS+RPC框架", "zmq_poll timeout / send-recv fail / try_again", "-", "send/recv/zmq_poll", "client1->worker1 控制面RPC", "优先排查socket读写失败与超时参数"),
    ("D03", "OS", "Get mmap entry failed / Pass fd… / invalid fd", "-", "mmap / recvmsg / sendmsg", "client1 mmapManager; fd_pass", "对齐同机SHM与fd传递时序"),
    ("D04", "OS", "K_OUT_OF_MEMORY(6); UB pool mmap", "-", "mmap(ANON)", "urma_manager InitMemoryBufferPool", "调大内存或降并发"),
    ("D05", "URMA", "1004; Failed to urma init/create/poll…", "urma_init, urma_create_*, urma_poll_jfc", "-", "urma_manager", "UMDK+设备+驱动"),
    ("D06", "URMA", "Failed to advise/import jfr", "urma_import_jfr, urma_advise_jfr", "-", "ExchangeJfr", "握手与对端 jfr 一致性"),
    ("D07", "URMA(降级)", "fallback to TCP/IP payload", "池分配失败链", "-", "PrepareUrmaBuffer", "性能问题查UB；功能成功可先记观测"),
    ("D08", "URMA", "1006; need to reconnect", "逻辑非单API", "-", "CheckUrmaConnectionStable", "重建 UB 会话"),
    ("D09", "数据系统逻辑", "etcd is unavailable", "-", "etcd访问syscall间接", "worker1 etcdStore", "etcd 续约与集群健康"),
    ("D10", "数据系统逻辑", "Query from master failed", "-", "-", "worker1->worker2", "Directory 与 hash ring 路由"),
    ("D11", "数据系统逻辑", "Get from remote failed; K_NOT_FOUND", "-", "-", "worker1->worker3", "副本与 Directory 目录一致性"),
    ("D12", "数据系统逻辑", "UB payload overflow / 响应计数不匹配", "-", "-", "client1 FillUrmaBuffer/GetBuffers", "版本/协议与 worker 返回一致性"),
    ("D13", "数据系统逻辑", "K_SCALING(32); scaling 文案", "-", "-", "MultiPublish worker内", "扩缩容窗口内重试或降载"),
    ("D14", "RPC框架+数据系统", "last_rc 触发 RetryOnError", "-", "-", "client1+worker1", "结合两端日志拆是传输还是业务码"),
]

URMA_ERR_HEADERS = [
    "URMA接口",
    "原始出错返回值(URMA枚举/返回)",
    "OS errno(编号)",
    "关联UDMA接口",
    "UDMA返回值/故障点",
    "错误原因(简述)",
    "典型日志片段",
    "备注",
]

URMA_ERR_ROWS = [
    ("[来自 umdk/urma_opcode.h]", "URMA_SUCCESS=0", "N/A", "-", "-", "调用成功", "-", "头文件确认"),
    ("[来自 umdk/urma_opcode.h]", "URMA_EAGAIN=EAGAIN", "EAGAIN(11)", "-", "-", "资源暂不可用，可重试", "-", "Resource temporarily unavailable"),
    ("[来自 umdk/urma_opcode.h]", "URMA_ENOMEM=ENOMEM", "ENOMEM(12)", "-", "-", "内存分配失败", "-", "Failed to allocate memory"),
    ("[来自 umdk/urma_opcode.h]", "URMA_ENOPERM=EPERM", "EPERM(1)", "-", "-", "操作不允许/权限不足", "-", "Operation not permitted"),
    ("[来自 umdk/urma_opcode.h]", "URMA_ETIMEOUT=ETIMEDOUT", "ETIMEDOUT(110)", "-", "-", "操作超时", "-", "Operation time out"),
    ("[来自 umdk/urma_opcode.h]", "URMA_EINVAL=EINVAL", "EINVAL(22)", "-", "-", "入参非法", "-", "Invalid argument"),
    ("[来自 umdk/urma_opcode.h]", "URMA_EEXIST=EEXIST", "EEXIST(17)", "-", "-", "对象已存在/重复初始化", "-", "Exist"),
    ("[来自 umdk/urma_opcode.h]", "URMA_EINPROGRESS=EINPROGRESS", "EINPROGRESS(115)", "-", "-", "异步处理中", "-", "通常等待后续完成"),
    ("[来自 umdk/urma_opcode.h]", "URMA_FAIL=0x1000", "N/A", "-", "-", "URMA通用失败码（非POSIX errno）", "-", "需结合接口上下文和日志"),
    ("urma_register_log_func", "URMA_EINVAL", "N/A", "-", "-", "参数校验失败（注册函数空指针）", "Invalid parameter", "来源于你给的清单"),
    ("urma_init", "URMA_EEXIST", "N/A", "-", "-", "重复初始化", "urma_init has been called before", "可通过初始化时序规避"),
    ("urma_init", "URMA_FAIL", "N/A", "-", "-", "provider/驱动so加载失败", "None of the providers registered", "检查UMDK与驱动部署"),
    ("urma_get_device_list", "nullptr", "EINVAL(22)", "-", "-", "参数为空或非法", "Invalid parameter", "接口入参检查"),
    ("urma_get_device_list", "nullptr", "ENODEV(19)", "-", "-", "设备数量为0", "-", "设备未识别或不可用"),
    ("urma_get_device_list", "nullptr", "ENOMEM(12)", "-", "-", "内存分配失败", "-", "系统内存不足"),
    ("urma_get_device_list", "nullptr", "ENOEXEC(8)", "-", "-", "设备数量前后不一致", "-", "异常场景，建议保留现场"),
    ("urma_get_eid_list", "null", "EINVAL(22)", "-", "-", "参数非法/空指针或eid规格异常", "invalid parameter with null_ptr", "含规格异常场景"),
    ("urma_get_eid_list", "null", "ENOMEM(12)", "-", "-", "OS内存分配失败", "-", "系统资源不足"),
    ("urma_get_eid_list", "null", "EIO(5)", "-", "-", "ioctl或内核侧EID信息缺失", "ioctl failed", "检查内核FE状态"),
    ("urma_register_seg", "null", "EINVAL(22)/N/A", "udma_u_alloc_tid", "NULL", "seg access/token参数非法或token分配失败", "Invalid parameter", "本地段权限组合非法"),
    ("urma_register_seg", "null", "N/A", "udma_u_register_seg", "NULL", "udma注册段失败（grant/cmd失败）", "-", "检查token_policy/access"),
    ("urma_query_device", "URMA_EINVAL", "EINVAL(22)", "udma_query_device", "NA", "参数错误", "Invalid parameter", "设备属性查询入参错误"),
    ("urma_query_device", "URMA_FAIL", "N/A", "udma_query_device", "NA", "底层udma查询失败", "Failed to query device attr", "常见于设备状态异常"),
    ("urma_str_to_eid", "-EINVAL", "N/A", "-", "-", "字符串格式错误或空指针", "format error", "输入EID格式需校验"),
    ("urma_create_context", "null", "EINVAL(22)/EIO(5)", "udma_u_create_context", "NULL/0", "参数非法/读取sysfs失败/设备打开失败", "Failed to open urma cdev", "可能见到db_addr mmap失败22"),
    ("urma_create_jfc", "null", "EINVAL(22)/N/A", "udma_u_create_jfc", "NULL/22/14", "参数非法、CQ buf分配失败、cmd失败", "jfc cfg depth of range", "深度和ceqn需合法"),
    ("urma_create_jfce", "null", "EINVAL(22)/N/A", "udma_u_create_jfse", "NULL", "参数非法、SQ buf分配失败、cmd失败", "-", "检查jfs配置"),
    ("urma_create_jfr", "null", "EINVAL(22)/ENOMEM(12)/N/A", "udma_u_create_jfr", "NULL", "参数非法、RQ/index buf分配失败、cmd失败", "-", "hugepage与mmap分配失败常见"),
    ("urma_create_jetty", "null", "EINVAL(22)/EPERM(1)/N/A", "udma_u_create_jetty", "NULL/22", "参数非法、group权限失败、cmd失败", "-", "RC模式/权限需匹配"),
    ("urma_import_jetty", "null", "EINVAL(22)/N/A", "udma_u_ctrlq_get_tp_list / udma_u_import_jetty_ex", "urma返回值/NULL", "参数非法、TP交换失败、导入失败", "-", "跨端会话参数需一致"),
    ("urma_bind_jetty", "URMA_EINVAL", "N/A", "udma_u_bind_jetty_ex", "22/17", "参数错误（trans_mode/tpn/tjetty）", "-", "协议与对象需匹配"),
    ("urma_bind_jetty", "URMA_ENOPERM", "N/A", "-", "-", "不是RC模式或序要求不匹配", "-", "需校验传输模式"),
    ("urma_bind_jetty", "URMA_FAIL/UDMA返回值", "N/A", "udma_u_ctrlq_get_tp_list / udma_u_bind_jetty_ex", "见故障点详细值", "底层udma或TP获取失败", "-", "可能看到URMA_FAIL(0x1000)"),
    ("urma_import_seg", "nullptr", "EINVAL(22)/N/A", "udma_u_import_seg", "NULL", "参数非法或target seg申请失败", "-", "检查token_policy与对端段信息"),
    ("urma_poll_jfc", "-1", "N/A", "udma_u_poll_jfc", "UDMA_INTER_ERR(1)", "参数错误或poll失败", "JFC_EMPTY(1)/JFC_POLL_ERR(2)", "与wait/ack/rearm联动定位"),
    ("urma_wait_jfc / urma_ack_jfc / urma_rearm_jfc", "失败返回", "N/A", "udma_u_poll_jfc(关联)", "UDMA_INTER_ERR(1)", "CQ事件流转失败", "Failed to wait/poll/rearm jfc", "常映射到1004"),
    ("urma_post_jetty_send_wr", "URMA_EINVAL/UDMA返回值", "N/A", "udma_u_post_jetty_send_wr", "URMA_EINVAL(22)", "参数错误(sge/opcode)或sqe设置失败", "-", "检查WR/SGE/远端信息"),
]


def style_header(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def style_body(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 12
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                max_len = max(max_len, min(56, len(str(v))))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def render_tree_chain(chain):
    """
    Render A|_B|_C into a readable multiline tree for Excel (forward call chain).

    Root at line 1; each deeper callee uses '└─' and 3-space indent per level
    so the column reads as an actual tree, not a flat |_ list.
    """
    if not isinstance(chain, str):
        return chain
    text = chain.strip()
    if "|_" not in text:
        return text
    parts = [p.strip() for p in text.split("|_") if p.strip()]
    if not parts:
        return chain
    if len(parts) == 1:
        return parts[0]
    lines = [parts[0]]
    for depth in range(1, len(parts)):
        indent = "   " * (depth - 1)
        lines.append(f"{indent}└─ {parts[depth]}")
    return "\n".join(lines)


def enrich_chain(chain, location, owner):
    """Add explicit RPC/user-validation nodes to chain text."""
    if not isinstance(chain, str):
        return chain
    out = chain
    loc = str(location).lower()
    own = str(owner)

    # Add component hint on root function: client/worker1/worker2/worker3
    comp_hint = "client组件"
    if "worker2" in loc:
        comp_hint = "worker-Directory查询"
    elif "worker3" in loc:
        comp_hint = "worker-数据拉取"
    elif "worker1" in loc:
        comp_hint = "worker组件"
    root, *rest = [p.strip() for p in out.split("|_")]
    if f"({comp_hint})" not in root:
        root = f"{root} ({comp_hint})"
    out = "|_".join([root] + rest) if rest else root

    if "->" in loc and "rpc" in loc and "stub_->" not in out and "retryonerror" not in out.lower():
        out = f"{out}|_ RPC调用"
    if "用户参数" in own and "k_invalid" not in out.lower() and "validate" not in out.lower():
        out = f"{out}|_ValidateRequest(K_INVALID)"
    return out


def split_location_detail(location):
    """Split location as [发生位置] + [详细描述]."""
    if not isinstance(location, str):
        return "-", "-"
    text = location.strip()
    if "（" in text and "）" in text:
        main = text.split("（", 1)[0].strip()
        detail = text.split("（", 1)[1].rsplit("）", 1)[0].strip()
    else:
        main, detail = text, "-"
    return main, detail


def derive_internal_error_enum_info(row):
    """Data-system error enum/value and meaning for Sheet1."""
    if len(row) < 7:
        return "-"
    status = str(row[6])
    text = f"{row[2]} {row[3]} {row[4]} {row[5]} {row[6]}".lower()
    lines = []

    def add(code, meaning):
        lines.append(f"- {code}: {meaning}")

    if "k_invalid" in status.lower() or "k_invalid" in text:
        add("K_INVALID", "入参非法（如 key/offset/timeout/batch 越界）")
    if "1001" in status or "timeout" in text:
        add("K_RPC_DEADLINE_EXCEEDED(1001)", "RPC超时，通常为网络抖动/队列等待/对端慢")
    if "1002" in status or "unavailable" in text:
        add("K_RPC_UNAVAILABLE(1002)", "RPC不可达，常见于连接失败或对端不可用")
    if "19" in status or "try_again" in text:
        add("K_TRY_AGAIN(19)", "瞬时错误，建议按策略重试")
    if "k_urma_error" in text or "1004" in status:
        add("K_URMA_ERROR(1004)", "URMA调用失败（初始化/队列/数据面）")
    if "1006" in status or "need to reconnect" in text:
        add("K_URMA_NEED_CONNECT(1006)", "URMA连接不稳定或已断开，需要重连")
    if "k_runtime_error" in status.lower() or "(5)" in status or "overflow" in text:
        add("K_RUNTIME_ERROR(5)", "运行时错误（如payload组装、mmap链路、URMA读写失败）")
    if "k_out_of_memory" in status.lower() or "(6)" in status or "oom" in text:
        add("K_OUT_OF_MEMORY(6)", "内存不足（UB池/请求处理过程）")
    if "k_scaling" in status.lower() or "32" in status:
        add("K_SCALING(32)", "集群扩缩容窗口触发的暂时性错误")
    if "k_not_found" in status.lower():
        add("K_NOT_FOUND", "对象不存在或已过期")
    if "k_unknown_error" in status.lower() or "invalid fd" in text:
        add("K_UNKNOWN_ERROR", "未知错误（常见于fd传递异常等）")

    if not lines:
        add("（无明确内部枚举）", "当前case以系统侧日志/状态为主，请结合Sheet2/3定位")
    return "\n".join(lines)


def normalize_status_with_enum(status_text):
    """Prefer enum(value) over plain numeric values in status text."""
    s = str(status_text)
    replacements = [
        ("1006", "K_URMA_NEED_CONNECT(1006)"),
        ("1004", "K_URMA_ERROR(1004)"),
        ("1002", "K_RPC_UNAVAILABLE(1002)"),
        ("1001", "K_RPC_DEADLINE_EXCEEDED(1001)"),
        ("32", "K_SCALING(32)"),
        ("19", "K_TRY_AGAIN(19)"),
        ("6", "K_OUT_OF_MEMORY(6)"),
        ("5", "K_RUNTIME_ERROR(5)"),
    ]
    for raw, enumv in replacements:
        s = s.replace(f"; {raw}", f"; {enumv}")
        s = s.replace(f"/{raw}", f"/{enumv}")
        s = s.replace(f" {raw};", f" {enumv};")
        s = s.replace(f" {raw},", f" {enumv},")
        if s.strip() == raw:
            s = enumv
    return s


def simplify_owner(owner, info, step):
    """Simplify owner categories per user rule: 用户参数/数据系统/OS/URMA."""
    o = str(owner)
    t = f"{owner} {info} {step}".lower()
    if "用户参数" in o:
        return "用户参数"
    if "urma" in t or "jfc" in t or "jfr" in t or "jfs" in t or "umdk" in t:
        return "URMA"
    # ZMQ-RPC framework internal logic (not direct socket/errno failure) -> 数据系统
    if "zmq" in t and not any(k in t for k in ["socket", "send", "recv", "connect", "poll", "errno", "eagain", "etimedout"]):
        return "数据系统"
    # RPC transport/socket/errno issues -> OS
    if "rpc" in t and any(k in t for k in ["socket", "send", "recv", "connect", "poll", "zmq", "errno", "timeout", "unavailable", "try_again"]):
        return "OS"
    if any(k in t for k in ["socket", "zmq", "send", "recv", "connect", "poll", "mmap", "fd", "errno", "close"]):
        return "OS"
    return "数据系统"


def build_chain_rows_with_system_info(rows):
    out = []
    for row_idx, row in enumerate(rows):
        # raw row layout:
        # 0接口 1调用链 2步骤 3发生位置 4责任 5接口信息 6典型status日志 7返回 8重试 9建议 10锚点 [11故障预期]
        if len(row) >= 12:
            iface, chain, step, location, owner, info, status_log, ret, retry, advice, anchor, fault_expect = row[:12]
        else:
            iface, chain, step, location, owner, info, status_log, ret, retry, advice, anchor = row[:11]
            fault_expect = "可失败（本行专述异常或边界条件）"
        row_core = (
            iface,
            chain,
            step,
            location,
            owner,
            info,
            status_log,
            ret,
            retry,
            advice,
            anchor,
        )
        chain = enrich_chain(chain, location, owner)
        loc_main, loc_detail = split_location_detail(location)
        tree_rendered = render_tree_chain(chain)
        # 合并列：故障预期 + 逻辑说明 + 正向树 + case；定界逆向流程见 Sheet5 + workbook/README
        chain_with_meta = (
            f"【故障预期】{fault_expect}\n"
            "【调用链逻辑】自上而下 = 调用先后顺序；[条件→] 表示仅部分请求经过的分支\n"
            "【调用链树 · 正向】根 ──► 叶（沿下行展开）\n"
            f"{tree_rendered}\n"
            "────────\n"
            f"【本行 case / 子阶段】{step}\n"
            f"[发生位置] {loc_main}\n"
            f"[详细描述] {loc_detail}"
        )
        preset = SHEET1_URMA_OS[row_idx]
        primary, urma_iface_p, urma_err_p, os_iface_p, os_err_p, locate_guide = preset
        urma_info, urma_err, os_info, os_err = merge_exclusive(
            primary, urma_iface_p, urma_err_p, os_iface_p, os_err_p
        )
        internal_info = derive_internal_error_enum_info(list(row_core))
        combined_status = internal_info
        root_cause = locate_guide
        owner_simple = simplify_owner(owner, info, step)
        if retry == "视阶段重试":
            return_retry = "返回: 状态直接返回；重试: 连接阶段通常快速失败，业务RPC阶段按 RetryOnError 对 1001/1002/19 退避重试。"
        elif retry in ("无", "-", "一般无"):
            return_retry = f"返回: {ret}；重试: 通常不重试。"
        else:
            return_retry = f"返回: {ret}；重试: {retry}。"
        reason_and_advice = f"{root_cause}\n- 补充建议: {advice}"
        evidence = _find_code_evidence(anchor, status_log)
        remark = f"{return_retry}\n- 代码锚点: {anchor}\n- {evidence}"
        out.append(
            (
                iface,
                chain_with_meta,
                owner_simple,
                combined_status,
                urma_info,
                urma_err,
                os_info,
                os_err,
                reason_and_advice,
                remark,
            )
        )
    return out


def fill_sheet(ws, headers, rows, tree_col=None):
    ws.append(headers)
    for row in rows:
        row_list = list(row)
        if tree_col is not None and 1 <= tree_col <= len(row_list):
            row_list[tree_col - 1] = render_tree_chain(row_list[tree_col - 1])
        ws.append(row_list)
    style_header(ws)
    style_body(ws)
    autosize(ws)
    if tree_col is not None:
        # keep tree column readable without over-expanding others
        ws.column_dimensions[get_column_letter(tree_col)].width = 64


def tune_sheet1_tree_layout(ws, tree_col=2, min_row=2):
    """Widen tree column and row heights so multiline trees are visible in Excel."""
    letter = get_column_letter(tree_col)
    ws.column_dimensions[letter].width = 88
    for row_idx in range(min_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=tree_col)
        val = cell.value
        if isinstance(val, str) and val.strip():
            lines = val.count("\n") + 1
            # ~15 pt per line, cap for Excel limit
            ws.row_dimensions[row_idx].height = min(409, max(18, lines * 15))


def tune_sheet1_urma_os_columns(ws, first_col=5, last_col=8):
    """URMA/OS 四列放具体接口与日志，拉宽便于阅读。"""
    for col in range(first_col, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 64


def _anchor_file_candidates(anchor):
    files = []
    for token in str(anchor).split(";"):
        token = token.strip()
        m = re.search(r"([A-Za-z0-9_./-]+\.(?:cpp|cc|c|h|hpp))", token)
        if m:
            files.append(Path(m.group(1)).name)
    return files


def _status_keywords(status_log):
    parts = []
    for p in str(status_log).split(";"):
        p = p.strip()
        if not p:
            continue
        if re.fullmatch(r"[A-Za-z_()]+", p):
            continue
        if re.fullmatch(r"[0-9/ ]+", p):
            continue
        if len(p) >= 6:
            parts.append(p)
    if not parts and status_log:
        parts.append(str(status_log))
    return parts[:2]


def _find_code_evidence(anchor, status_log):
    """Find an in-code snippet by anchor files + log keyword."""
    if not SOURCE_ROOT.exists():
        return "代码证据: 源码目录不存在"
    file_names = _anchor_file_candidates(anchor)
    if not file_names:
        return f"代码证据: 锚点 {anchor}"
    keywords = _status_keywords(status_log)
    for name in file_names:
        for path in SOURCE_ROOT.rglob(name):
            try:
                lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
            except Exception:
                continue
            for kw in keywords:
                kw_lower = kw.lower()
                for i, ln in enumerate(lines):
                    if kw_lower in ln.lower():
                        start = max(0, i - 1)
                        end = min(len(lines), i + 2)
                        snippet = " | ".join(f"{idx+1}:{lines[idx].strip()}" for idx in range(start, end))
                        return f"代码证据: {path.name}:{i+1} {snippet}"
            # fallback: use first line containing a common log/error call
            for i, ln in enumerate(lines):
                if any(tok in ln for tok in ["LOG_", "RETURN_STATUS", "CHECK_FAIL_RETURN_STATUS", "FormatString("]):
                    return f"代码证据: {path.name}:{i+1} {i+1}:{ln.strip()}"
    return f"代码证据: 未命中关键词，锚点 {anchor}"


def main():
    wb = Workbook()
    ws_chain = wb.active
    ws_chain.title = "Sheet1_调用链路分析"
    chain_rows = build_chain_rows_with_system_info(CHAIN_ROWS)
    fill_sheet(ws_chain, CHAIN_HEADERS, chain_rows, tree_col=None)
    tune_sheet1_tree_layout(ws_chain, tree_col=2)
    tune_sheet1_urma_os_columns(ws_chain, 5, 8)
    hdr_b = ws_chain.cell(row=1, column=2)
    hdr_b.comment = Comment(
        "本列结构：【故障预期】无错误=健康主路径各段假定不发生故障；其余行多为「可失败」专述异常。"
        "【调用链逻辑】+【调用链树】自上而下为调用顺序，|_ 在数据中行内渲染为 └─ 树。"
        "E~H 列：URMA 与 OS **错误列互斥**（见 sheet1_system_presets.merge_exclusive）；"
        "具体接口与日志原文见同表 sheet1_system_presets.py。"
        "逆向定界：先 Status/日志 → Trace → Sheet5；见 workbook/README。",
        "generate_kv_client_observability_xlsx",
    )

    ws_os = wb.create_sheet("Sheet2_OS系统调用查表")
    fill_sheet(ws_os, OS_HEADERS, OS_ROWS, tree_col=2)

    ws_urma = wb.create_sheet("Sheet3_URMA接口查表")
    fill_sheet(ws_urma, URMA_HEADERS, URMA_ROWS, tree_col=2)

    ws_perf = wb.create_sheet("Sheet4_性能关键路径")
    fill_sheet(ws_perf, PERF_HEADERS, PERF_ROWS)

    ws_dem = wb.create_sheet("Sheet5_定界-case查表")
    fill_sheet(ws_dem, DEMARCATION_HEADERS, DEMARCATION_ROWS)

    ws_urma_err = wb.create_sheet("Sheet6_URMA错误码解释")
    fill_sheet(ws_urma_err, URMA_ERR_HEADERS, URMA_ERR_ROWS)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
