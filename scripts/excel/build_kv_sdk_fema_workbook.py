#!/usr/bin/env python3
"""Build KV SDK (Init / MCreate / MSet / MGet) FMEA + observability Excel workbook.

Reads: workspace/reliability/kv_sdk_fema_rows.tsv
Writes: workspace/reliability/kv_sdk_fema_analysis.xlsx

Requires: openpyxl
Run from vibe-coding-files repo root:
  python3 scripts/excel/build_kv_sdk_fema_workbook.py
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
TSV_PATH = REPO_ROOT / "workspace" / "reliability" / "kv_sdk_fema_rows.tsv"
OUT_XLSX = REPO_ROOT / "workspace" / "reliability" / "kv_sdk_fema_analysis.xlsx"

WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_FONT = Font(bold=True)

STATUS_DRILLDOWN = [
    [
        "StatusCode",
        "枚举名",
        "第一步问什么",
        "日志关键词(与Trace同屏)",
        "URMA专项(同TraceID检索)",
        "下一步定界文档",
    ],
    [
        "2",
        "K_INVALID",
        "是否参数/batch/地址配置错误？",
        "Invalid|CheckFail|Validate",
        "—",
        "docs/observable/kv-client-excel/kv-client-Sheet1",
    ],
    [
        "3",
        "K_NOT_FOUND",
        "key 是否真存在？TTL？",
        "NOT_FOUND|Key not found",
        "—",
        "docs/observable/kv-client-Get路径-树状错误矩阵.md",
    ],
    [
        "5",
        "K_RUNTIME_ERROR",
        "是否 executor/内部异常？是否伴随 OOM/IO？",
        "Runtime error|exception",
        "若发生在 UB 路径，同 Trace 搜 URMA",
        "docs/reliability/client-status-codes-evidence-chain.md",
    ],
    [
        "6",
        "K_OUT_OF_MEMORY",
        "节点/容器内存与 cgroup？",
        "Out of memory|OOM",
        "—",
        "00-kv-client-fema-scenarios-failure-modes.md",
    ],
    [
        "7",
        "K_IO_ERROR",
        "磁盘/shm/权限？",
        "IO error|pread|pwrite|mmap",
        "—",
        "00-kv-client-fema-read-paths-reliability.md",
    ],
    [
        "8",
        "K_NOT_READY",
        "Init 是否完成？是否并发 Init？",
        "Not ready|ProcessInit",
        "Init 阶段 URMA 失败可导致未就绪",
        "kv-client-URMA-OS-读写初始化-跨模块错误与重试.md",
    ],
    [
        "13",
        "K_NO_SPACE",
        "磁盘/shm 满？",
        "No space|spill",
        "—",
        "00-kv-client-fema-scenarios-failure-modes.md",
    ],
    [
        "18",
        "K_FILE_LIMIT_REACHED",
        "ulimit -n / 容器 FD？",
        "FD|file limit",
        "—",
        "00-kv-client-visible-status-codes.md",
    ],
    [
        "19",
        "K_TRY_AGAIN",
        "瞬时故障？与 1008/1001 同时出现？",
        "Try again|retry",
        "UB 闪断常见",
        "00-kv-client-visible-status-codes.md §2",
    ],
    [
        "23",
        "K_CLIENT_WORKER_DISCONNECT",
        "Worker 是否重启/网络断？",
        "Disconnected|reconnect|heartbeat",
        "重连后 URMA 需重建",
        "kv-client-Sheet3-TCP-RPC对照.md",
    ],
    [
        "25",
        "K_MASTER_TIMEOUT",
        "etcd/Master 延迟？",
        "Master|etcd|timeout",
        "—",
        "00-kv-client-fema-scenarios-failure-modes.md §etcd",
    ],
    [
        "30-32",
        "K_RETRY_IF_LEAVING / K_SCALE_DOWN / K_SCALING",
        "是否在缩容/迁移窗口？",
        "scale|leaving|scaling",
        "—",
        "worker-续约-健康检查-资源-故障检测与告警.md",
    ],
    [
        "1001",
        "K_RPC_DEADLINE_EXCEEDED",
        "超时阈值 vs p99？仅 UB 还是 TCP 也慢？",
        "deadline|timeout|RPC",
        "同 Trace 看 URMA_WAIT 与 RPC 先后",
        "get-latency-timeout-sensitive-analysis-5ms-20ms.md",
    ],
    [
        "1002",
        "K_RPC_UNAVAILABLE",
        "对端不可达？连接数？",
        "unavailable|channel",
        "—",
        "kv-client-Sheet3-TCP-RPC对照.md",
    ],
    [
        "1004",
        "K_URMA_ERROR",
        "UMDK 返回？设备状态？",
        "URMA_ERROR|urma_|status",
        "**同 TraceID 全 URMA 日志**",
        "kv-client-URMA-错误枚举与日志-代码证据.md",
    ],
    [
        "1006",
        "K_URMA_NEED_CONNECT",
        "连接是否被 tear down？",
        "NEED_CONNECT|reconnect",
        "握手与 IMPORT 段",
        "kv-client-URMA-OS-读写初始化-跨模块错误与重试.md",
    ],
    [
        "1008",
        "K_URMA_TRY_AGAIN",
        "UB 闪断/重传？",
        "TRY_AGAIN|poll jfc",
        "jfc|event wait 与 1008 对齐看",
        "kv-client-URMA-错误枚举与日志-代码证据.md",
    ],
    [
        "2000",
        "K_OC_ALREADY_SEALED",
        "是否重复 Publish？",
        "sealed|Seal",
        "—",
        "kv-client-Sheet1",
    ],
    [
        "2003",
        "K_WRITE_BACK_QUEUE_FULL",
        "写回队列是否打满？",
        "queue full|write back",
        "—",
        "kv-client-写接口-定位定界.md",
    ],
    [
        "2004",
        "K_OC_KEY_ALREADY_EXIST",
        "NX 语义？",
        "already exist",
        "—",
        "—",
    ],
]

FOOL_GUIDE = [
    ["步骤", "你要做的事", "产出物/判据"],
    [
        "1",
        "从 SDK 返回值记下 StatusCode 数值与 ToString() 全文",
        "一行原始错误文本",
    ],
    [
        "2",
        "在 client 日志中按 **同一 TraceID**（或时间窗 ±1s）拉出相邻行",
        "带 UUID 的日志片段",
    ],
    [
        "3",
        "若码在 1004/1006/1008 或日志含 urma_|jfc|IMPORT：打开 Sheet「StatusCode_抽丝剥茧」URMA 列 + observable URMA 证据 md",
        "是否 UB 面故障初判",
    ],
    [
        "4",
        "若码在 1001/1002/23：打开 TCP-RPC 对照与 ZMQ 段，查 connect/heartbeat",
        "是否 TCP/RPC 面",
    ],
    [
        "5",
        "若码在 25/14/19 且伴 etcd：对齐 Master/etcd 监控与 00-FEMA etcd 章节",
        "是否控制面",
    ],
    [
        "6",
        "在 Sheet「FMEA_Init_MCreate_MSet_MGet」按 API 列筛选你的接口行，读检测与恢复列",
        "匹配到的 FM_ID",
    ],
    [
        "7",
        "需要性能关联时：对齐全局 Perf 日志与 docs 中 PERF_KEY（见 log-perf-zmq-kv-observability-design）",
        "是否慢在 RPC 还是 URMA 等待",
    ],
    [
        "8",
        "仍不清：用 kv-client-观测 xlsx Sheet5 定界-case 表做最后一跳",
        "责任域：用户参数/OS/URMA/RPC/逻辑",
    ],
]


def load_tsv(path: Path) -> tuple[list[str], list[list[str]]]:
    if not path.exists():
        raise FileNotFoundError(path)
    with path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f, delimiter="\t"))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def autosize(ws, max_rows: int = 600, max_width: int = 52) -> None:
    col_count = ws.max_column
    for col_idx in range(1, col_count + 1):
        w = 10
        for r in ws.iter_rows(
            min_row=1, max_row=min(ws.max_row, max_rows), min_col=col_idx, max_col=col_idx
        ):
            for cell in r:
                if cell.value is not None:
                    w = max(w, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def write_table(ws, header: list[str], rows: list[list[str]], bold_header: bool = True) -> None:
    ws.append(header)
    if bold_header:
        for c in range(1, len(header) + 1):
            ws.cell(row=1, column=c).font = HEADER_FONT
            ws.cell(row=1, column=c).alignment = WRAP
    for row in rows:
        ws.append(row)
        for c in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=c).alignment = WRAP


def main() -> int:
    header, data = load_tsv(TSV_PATH)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "README_使用说明"
    readme_lines = [
        "KV SDK FMEA + 定位定界联动工作簿",
        "",
        "数据源: workspace/reliability/kv_sdk_fema_rows.tsv（可编辑后重新运行脚本生成）",
        "生成命令: python3 scripts/excel/build_kv_sdk_fema_workbook.py",
        "",
        "Sheet 说明:",
        "  - FMEA_Init_MCreate_MSet_MGet: 故障模式 / 检测 / 恢复 / 与 observable·reliability 文档锚点",
        "  - StatusCode_抽丝剥茧: 从返回码下钻的问题树 + URMA+Trace 检索提示",
        "  - 傻瓜步骤: 一线按表操作顺序",
        "  - Mermaid_流程图源码: 粘贴到支持 Mermaid 的编辑器预览",
        "",
        "URMA 检测原则: 必须使用与业务日志相同的 TraceID 关联检索 URMA 相关行（见 observable kv-client-excel）",
        "",
        "权威 StatusCode: yuanrong-datasystem include/datasystem/utils/status.h + status_code.def",
    ]
    for i, line in enumerate(readme_lines, start=1):
        ws0.cell(row=i, column=1, value=line)
    ws0.column_dimensions["A"].width = 96

    ws1 = wb.create_sheet("FMEA_Init_MCreate_MSet_MGet")
    write_table(ws1, header, data)
    autosize(ws1)

    ws2 = wb.create_sheet("StatusCode_抽丝剥茧")
    write_table(ws2, STATUS_DRILLDOWN[0], STATUS_DRILLDOWN[1:])
    autosize(ws2)

    ws3 = wb.create_sheet("傻瓜步骤")
    write_table(ws3, FOOL_GUIDE[0], FOOL_GUIDE[1:])
    autosize(ws3)

    mermaid = """flowchart TD
    A[SDK 返回非 OK] --> B{记下 Code + TraceID}
    B --> C{1004/1006/1008 或 日志含 urma_?}
    C -->|是| D[同 TraceID 拉 URMA 日志 + Sheet2 URMA列]
    C -->|否| E{1001/1002/23/29?}
    E -->|是| F[TCP-RPC 对照 + heartbeat/reconnect]
    E -->|否| G{25/14/19 + etcd?}
    G -->|是| H[Master/etcd FEMA + 监控]
    G -->|否| I[按 API 筛 FMEA 表 + Sheet5 定界-case]
    D --> J[恢复: UB/驱动/降级验证]
    F --> J
    H --> J
    I --> J
"""
    ws4 = wb.create_sheet("Mermaid_流程图源码")
    ws4.cell(row=1, column=1, value="将下列源码复制到 Mermaid Live / Typora / VS Code 插件预览:")
    ws4.cell(row=2, column=1, value=mermaid.strip())
    ws4["A1"].font = HEADER_FONT
    ws4["A2"].alignment = WRAP
    ws4.column_dimensions["A"].width = 88

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
